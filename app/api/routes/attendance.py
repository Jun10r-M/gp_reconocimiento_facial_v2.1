from fastapi import APIRouter, HTTPException, Depends, status, Header, Response
import csv
import io
from app.services.auth_service import get_current_admin, require_permission
from app.models.attendance import ShiftCreate, ShiftResponse, AttendanceLogResponse, BulkManualPunchRequest
from app.services.attendance_service import AttendanceService
from app.services.employee_service import EmployeeService
from app.repositories.shift_repository import ShiftRepository
from typing import List, Optional

router = APIRouter(tags=["Attendance"])

def get_attendance_service() -> AttendanceService:
    return AttendanceService()

def get_employee_service() -> EmployeeService:
    return EmployeeService()

def get_shift_repository() -> ShiftRepository:
    return ShiftRepository()

@router.get(
    "/attendance/shifts",
    summary="Get paginated weekly matrix of employee shifts",
    description="Returns a paginated list of employees and their scheduled weekly shifts, avoiding N+1 client queries.",
    tags=["Shifts"]
)
async def get_paginated_shifts(
    page: Optional[int] = None,
    limit: Optional[int] = 10,
    search: Optional[str] = None,
    attendance_service: AttendanceService = Depends(get_attendance_service),
    employee_service: EmployeeService = Depends(get_employee_service),
    shift_repo: ShiftRepository = Depends(get_shift_repository),
    admin_session: dict = Depends(require_permission("shifts", "read"))
):
    if page is None:
        employees = await employee_service.get_all_employees()
        items = []
        for emp in employees:
            shifts = await shift_repo.get_by_employee_id(emp["id"])
            items.append({
                "employee": emp,
                "shifts": [
                    {
                        "id": s["id"],
                        "employee_id": s["employee_id"],
                        "day_of_week": s["day_of_week"],
                        "start_time": str(s["start_time"]),
                        "end_time": str(s["end_time"]),
                        "tolerance": s.get("tolerance") if s.get("tolerance") is not None else 10
                    } for s in shifts
                ]
            })
        return items

    # Paginar colaboradores
    emp_pag = await employee_service.get_paginated_employees(page, limit, search)
    
    items = []
    for emp in emp_pag["items"]:
        shifts = await shift_repo.get_by_employee_id(emp["id"])
        items.append({
            "employee": emp,
            "shifts": [
                {
                    "id": s["id"],
                    "employee_id": s["employee_id"],
                    "day_of_week": s["day_of_week"],
                    "start_time": str(s["start_time"]),
                    "end_time": str(s["end_time"]),
                    "tolerance": s.get("tolerance") if s.get("tolerance") is not None else 10
                } for s in shifts
            ]
        })

    return {
        "items": items,
        "total": emp_pag["total"],
        "page": emp_pag["page"],
        "limit": emp_pag["limit"],
        "pages": emp_pag["pages"]
    }

@router.get(
    "/data",
    summary="Dashboard data compiler (CU-11)",
    description="Compiles and returns all employee profiles (with active positions/salaries) and chronological attendance logs for the dashboard."
)
async def get_dashboard_data(
    attendance_service: AttendanceService = Depends(get_attendance_service),
    employee_service: EmployeeService = Depends(get_employee_service),
    admin_session: dict = Depends(require_permission("attendance_logs", "read"))
):
    try:
        attendance_logs = await attendance_service.get_all_records()
        employees = await employee_service.get_all_employees()
        return {
            "attendance": attendance_logs,
            "employees": employees
        }
    except Exception as e:
        print(f"[API ERROR] Error compilando datos del dashboard: {e}")
        return {"attendance": [], "employees": []}

@router.get(
    "/attendance/latest-public",
    summary="Get latest public punch for scanner visual feedback",
    description="Returns the latest attendance punch (limit 1) without requiring authentication, to update the scanner UI."
)
async def get_latest_public_punch(
    attendance_service: AttendanceService = Depends(get_attendance_service),
    employee_service: EmployeeService = Depends(get_employee_service)
):
    try:
        logs = await attendance_service.repository.get_all()
        if not logs:
            return None
        
        # Filtrar logs válidos y ordenarlos por timestamp ascendente
        valid_logs = [l for l in logs if l.get("timestamp")]
        if not valid_logs:
            return None
            
        sorted_logs = sorted(valid_logs, key=lambda x: x["timestamp"])
        latest = sorted_logs[-1] # El último elemento es el más reciente

        emp = await employee_service.get_employee_by_id(latest["employee_id"])
        
        from datetime import datetime
        ts = latest["timestamp"]
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, datetime) else str(ts)
        
        return {
            "id": latest["id"],
            "employee_id": latest["employee_id"],
            "name": emp.get("full_name", "Empleado"),
            "documento": emp.get("document_number", emp.get("documento", "")),
            "puesto": emp.get("position", "Sin Cargo"),
            "time": ts_str,
            "action": "Marcación",
            "method": latest["method"]
        }
    except Exception as e:
        print(f"[API ERROR] Error obteniendo última marcación pública: {e}")
        return None

@router.get(
    "/attendance/logs",
    summary="Get biometric logs / Bitácora (CU-11)",
    description="Returns a chronological list of all biometric interactions, specifying whether facial recognition or fingerprint was used. Supports pagination with page, limit and search."
)
async def get_attendance_logs(
    page: Optional[int] = None,
    limit: Optional[int] = 10,
    search: Optional[str] = None,
    service: AttendanceService = Depends(get_attendance_service),
    admin_session: dict = Depends(require_permission("attendance_logs", "read"))
):
    if page is not None:
        return await service.get_paginated_records(page, limit, search)
    return await service.get_all_records()

@router.post(
    "/attendance/shifts",
    response_model=ShiftResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Assign shifts and schedule work hours (CU-07)",
    description="Schedules expected working hours for an employee on a specific day of the week.",
    tags=["Shifts"]
)
async def create_or_update_shift(
    shift: ShiftCreate, 
    repo: ShiftRepository = Depends(get_shift_repository),
    admin_session: dict = Depends(require_permission("shifts", "create"))
):
    shift_id = await repo.create_or_update({
        "employee_id": shift.employee_id,
        "day_of_week": shift.day_of_week,
        "start_time": shift.start_time,
        "end_time": shift.end_time,
        "tolerance": shift.tolerance if shift.tolerance is not None else 10
    })
    return {
        "id": shift_id,
        "employee_id": shift.employee_id,
        "day_of_week": shift.day_of_week,
        "start_time": shift.start_time,
        "end_time": shift.end_time,
        "tolerance": shift.tolerance if shift.tolerance is not None else 10
    }

@router.get(
    "/attendance/shifts/employee/{employee_id}",
    response_model=List[ShiftResponse],
    summary="Get employee schedules",
    description="Returns the planned work schedule for a worker across the week.",
    tags=["Shifts"]
)
async def get_shifts_by_employee(
    employee_id: int, 
    repo: ShiftRepository = Depends(get_shift_repository),
    admin_session: dict = Depends(require_permission("shifts", "read"))
):
    shifts = await repo.get_by_employee_id(employee_id)
    return [
        {
            "id": s["id"],
            "employee_id": s["employee_id"],
            "day_of_week": s["day_of_week"],
            "start_time": str(s["start_time"]),
            "end_time": str(s["end_time"]),
            "tolerance": s.get("tolerance") if s.get("tolerance") is not None else 10
        } for s in shifts
    ]

from pydantic import BaseModel

class ShiftUpdate(BaseModel):
    start_time: str
    end_time: str
    tolerance: Optional[int] = 10

@router.put(
    "/attendance/shifts/{shift_id}",
    response_model=ShiftResponse,
    summary="Update assigned shift hours",
    description="Updates start and end hours of a scheduled shift.",
    tags=["Shifts"]
)
async def update_shift(
    shift_id: int,
    data: ShiftUpdate,
    repo: ShiftRepository = Depends(get_shift_repository),
    admin_session: dict = Depends(require_permission("shifts", "update"))
):
    admin_user = admin_session["username"]
    shift = await repo.get_by_id(shift_id)
    if not shift:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Turno no encontrado."
        )

    tolerance_val = data.tolerance if data.tolerance is not None else 10
    await repo.update(
        shift_id=shift_id,
        start_time=data.start_time,
        end_time=data.end_time,
        tolerance=tolerance_val,
        updated_by=admin_user
    )

    # Log action to audit trail
    from app.services.audit_log_service import AuditLogService
    await AuditLogService().log_action(
        username=admin_user,
        action="EDITAR_TURNO",
        description=f"Edición de horario para el Turno ID: {shift_id} (Nuevo horario: {data.start_time} - {data.end_time}, Tolerancia: {tolerance_val}m) de Empleado ID: {shift['employee_id']}."
    )

    return {
        "id": shift_id,
        "employee_id": shift["employee_id"],
        "day_of_week": shift["day_of_week"],
        "start_time": data.start_time,
        "end_time": data.end_time,
        "tolerance": tolerance_val
    }

from app.services.biometrics_service import BiometricsService

def get_biometrics_service() -> BiometricsService:
    return BiometricsService()

@router.post(
    "/attendance/punch-fingerprint",
    summary="Register attendance using fingerprint fallback (CU-09)",
    description="Allows employees to clock in/out using their fingerprint if facial recognition is offline or fails."
)
async def punch_fingerprint(
    document_number: str,
    fingerprint_data: str,
    employee_service: EmployeeService = Depends(get_employee_service),
    biometrics_service: BiometricsService = Depends(get_biometrics_service),
    attendance_service: AttendanceService = Depends(get_attendance_service)
):
    emp = await employee_service.get_employee_by_document(document_number)
    if not emp:
        raise HTTPException(status_code=404, detail="Colaborador no encontrado con el DNI provisto.")
        
    employee_id = emp["id"]
    verified = await biometrics_service.verify_fingerprint(employee_id, fingerprint_data)
    if not verified:
        raise HTTPException(status_code=400, detail="Huella dactilar no coincide. Verificación rechazada.")
        
    action, timestamp = await attendance_service.register_punch(employee_id, "fingerprint")
    
    return {
        "status": "success",
        "name": emp["full_name"],
        "documento": emp["document_number"],
        "puesto": emp.get("position", "Sin Cargo"),
        "action": action,
        "time": timestamp,
        "method": "fingerprint"
    }

@router.post(
    "/attendance/punch-manual/bulk",
    summary="Register bulk manual punches (CU-10)",
    description="Allows administrators to log multiple manual clock-in/out records at once."
)
async def punch_manual_bulk(
    request: BulkManualPunchRequest,
    attendance_service: AttendanceService = Depends(get_attendance_service),
    admin_session: dict = Depends(require_permission("attendance_logs", "create"))
):
    admin_user = admin_session["username"]

    punches_dict = [{"employee_id": p.employee_id, "timestamp": p.timestamp} for p in request.punches]
    count = await attendance_service.register_bulk_manual_punches(punches_dict, admin_user)

    # Log action to audit trail
    from app.services.audit_log_service import AuditLogService
    await AuditLogService().log_action(
        username=admin_user,
        action="MARCACIÓN_MANUAL_MASIVA",
        description=f"Registro de {count} marcaciones manuales de contingencia en lote."
    )

    return {
        "status": "success",
        "message": f"Se registraron exitosamente {count} marcaciones manuales.",
        "count": count
    }

@router.get(
    "/attendance/export/employee/{employee_id}",
    summary="Export attendance logs for a specific employee as Excel or PDF",
    description="Generates a downloadable file (Excel .xlsx or PDF) with chronological attendance logs of a specific worker."
)
async def export_employee_attendance(
    employee_id: int,
    format: str = "excel",
    service: AttendanceService = Depends(get_attendance_service),
    employee_service: EmployeeService = Depends(get_employee_service),
    admin_session: dict = Depends(require_permission("attendance_logs", "read"))
):
    emp = await employee_service.get_employee_by_id(employee_id)
    if not emp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Colaborador no encontrado."
        )

    logs = await service.repository.get_by_employee_id(employee_id)

    from datetime import datetime as dt

    def fmt_ts(ts):
        return ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, dt) else str(ts)

    def method_label(m):
        if m == "fingerprint": return "Huella Digital"
        if m == "manual":      return "Marcación Manual"
        return "Reconocimiento Facial"

    clean_name = emp["full_name"].replace(" ", "_").lower()

    # ── EXCEL ──────────────────────────────────────────────────────────────
    if format.lower() != "pdf":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Paleta de colores
        COLOR_HEADER_BG  = "1E293B"   # slate-800
        COLOR_HEADER_FG  = "F8FAFC"   # casi blanco
        COLOR_ROW_ODD    = "F1F5F9"   # slate-100
        COLOR_ROW_EVEN   = "FFFFFF"

        # Encabezado del reporte (fila 1)
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Asistencia — {emp['full_name']}  |  DNI: {emp['document_number']}"
        title_cell.font      = Font(bold=True, size=13, color=COLOR_HEADER_FG)
        title_cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Fila 2: columnas
        headers = ["#", "Fecha y Hora", "Colaborador", "DNI", "Método de Marcación"]
        header_fill = PatternFill("solid", fgColor="0F172A")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font      = Font(bold=True, color=COLOR_HEADER_FG, size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Datos
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for i, log in enumerate(logs, 1):
            row = i + 2
            fill_color = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
            row_fill   = PatternFill("solid", fgColor=fill_color)

            values = [i, fmt_ts(log["timestamp"]), emp["full_name"], emp["document_number"], method_label(log["method"])]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.fill      = row_fill
                cell.border    = border
                cell.alignment = Alignment(horizontal="center" if col_idx in (1, 4) else "left", vertical="center")
            ws.row_dimensions[row].height = 18

        # Anchos de columna
        col_widths = [6, 22, 30, 14, 24]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)

        filename = f"asistencia_{clean_name}_{emp['document_number']}.xlsx"
        return Response(
            content=excel_buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ── PDF ────────────────────────────────────────────────────────────────
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    pdf_buf = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm
    )

    styles   = getSampleStyleSheet()
    dark_bg  = colors.HexColor("#1E293B")
    mid_bg   = colors.HexColor("#0F172A")
    odd_row  = colors.HexColor("#F1F5F9")
    even_row = colors.white
    accent   = colors.HexColor("#2DD4BF")

    title_style = ParagraphStyle(
        "titulo", parent=styles["Title"],
        fontSize=14, textColor=colors.white, alignment=TA_CENTER,
        spaceAfter=4
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#94A3B8"), alignment=TA_CENTER,
        spaceAfter=12
    )

    story = [
        Paragraph(f"Reporte de Asistencia", title_style),
        Paragraph(f"Colaborador: <b>{emp['full_name']}</b>  |  DNI: {emp['document_number']}  |  Total de registros: {len(logs)}", sub_style),
        Spacer(1, 0.3*cm),
    ]

    # Tabla
    col_headers = ["#", "Fecha y Hora", "Colaborador", "DNI", "Método de Marcación"]
    table_data = [col_headers]
    for i, log in enumerate(logs, 1):
        table_data.append([
            str(i),
            fmt_ts(log["timestamp"]),
            emp["full_name"],
            emp["document_number"],
            method_label(log["method"])
        ])

    col_widths_pdf = [1.2*cm, 5*cm, 7*cm, 3*cm, 5.5*cm]
    table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

    style_cmds = [
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0), mid_bg),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 9),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 8),
        # Data rows alternating
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("ALIGN",         (0, 1), (0, -1), "CENTER"),
        ("ALIGN",         (3, 1), (3, -1), "CENTER"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [even_row, odd_row]),
        # Grid
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("LINEBELOW",     (0, 0), (-1, 0), 1.5, accent),
    ]
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    doc.build(story)
    pdf_buf.seek(0)

    filename = f"asistencia_{clean_name}_{emp['document_number']}.pdf"
    return Response(
        content=pdf_buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

